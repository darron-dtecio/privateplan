"""Ingest a company-published IR financial workbook (.xlsx).

Company IR workbooks (e.g. Microsoft's FinancialStatementFY26Q4.xlsx) are fresher
than EDGAR XBRL right after earnings and are the practical source for segment
revenue. Layouts vary by company, so this module:

1. downloads the workbook and dumps every sheet as a trimmed grid to
   ir_workbook.json (so the analyst/Claude can inspect anything missed), and
2. best-effort auto-extracts quarterly segment revenue from sheets whose names
   mention segments.
"""

from __future__ import annotations

import io
import json
import re
from datetime import datetime

import openpyxl
import requests

import contact

HEADERS = {"User-Agent": contact.user_agent("IR workbook ingest")}

QUARTER_PAT = re.compile(r"(Q[1-4])\s*[- ]?\s*(FY)?\s*(\d{2,4})|\b(FY)?\s?(\d{2,4})\s*[- ]?(Q[1-4])", re.I)


def _cell_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def download_workbook(url: str) -> openpyxl.Workbook:
    resp = requests.get(url, headers=HEADERS, timeout=60)
    resp.raise_for_status()
    return openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True, read_only=True)


def workbook_to_grids(wb: openpyxl.Workbook, max_rows: int = 80, max_cols: int = 30) -> dict:
    """Dump each sheet to a trimmed grid of values."""
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        grid = []
        for r, row in enumerate(ws.iter_rows(values_only=True)):
            if r >= max_rows:
                break
            vals = [(_cell_str(v) if not isinstance(v, (int, float)) else v) for v in row[:max_cols]]
            while vals and (vals[-1] == "" or vals[-1] is None):
                vals.pop()
            grid.append(vals)
        while grid and not grid[-1]:
            grid.pop()
        sheets[name] = grid
    return sheets


def _quarter_columns(row: list) -> dict[int, str]:
    """Columns in a row whose text looks like a quarter label (Q1-25, FY26 Q1, ...)."""
    return {ci: val for ci, val in enumerate(row)
            if isinstance(val, str) and QUARTER_PAT.search(val)}


def extract_segments(grids: dict, sheet_hint: str = "HISTORY") -> list[dict]:
    """Best-effort extraction of quarterly line items from history-style sheets.

    Company workbooks (e.g. MSFT's 'Segment History' / 'Products & Services
    History') repeat blocks where the header row holds the block name in col 0
    and quarter labels across; data rows follow until the next header.
    """
    out = []
    for name, grid in grids.items():
        upper = name.upper()
        if sheet_hint not in upper and "SEGMENT" not in upper:
            continue
        qcols: dict[int, str] = {}
        section = None
        for row in grid:
            if not row:
                continue
            header_cols = _quarter_columns(row)
            if len(header_cols) >= 2:
                qcols = header_cols
                section = _cell_str(row[0]) or section
                continue
            if not qcols:
                continue
            label = _cell_str(row[0])
            numbers = {qcols[ci]: row[ci] for ci in qcols
                       if ci < len(row) and isinstance(row[ci], (int, float))}
            if label and numbers:
                out.append({
                    "sheet": name,
                    "section": section,
                    "line_item": label,
                    "values": numbers,
                })
            elif label and not numbers:
                section = label
    return out


def parse_ir_quarter(label: str) -> tuple[int, int] | None:
    m = re.fullmatch(r"Q([1-4])-(\d{2})", label.strip())
    if m:
        return 2000 + int(m.group(2)), int(m.group(1))
    return None


def to_segments(ir: dict) -> dict | None:
    """Normalize an ingested IR workbook into the segments.json shape.

    Returns {"source": "ir_workbook", "source_url", "segments": [{"name",
    "quarters": {"FY26 Q1": dollars}}]} or None when nothing usable was found.
    IR workbooks state values in millions; quarter keys become the pipeline's
    canonical fiscal labels.
    """
    seg_rows = [s for s in ir.get("segments", [])
                if s.get("line_item") == "Revenue" and s.get("section")
                and s["section"].lower() != "total" and "SEGMENT" in s["sheet"].upper()]
    merged: dict[str, dict[tuple[int, int], float]] = {}
    for s in seg_rows:
        tgt = merged.setdefault(s["section"], {})
        for k, v in s["values"].items():
            q = parse_ir_quarter(k)
            if q:
                tgt[q] = v * 1e6
    if not merged:
        return None
    return {
        "source": "ir_workbook",
        "source_url": ir.get("source_url"),
        "unit": "USD",
        "as_of": datetime.now().date().isoformat(),
        "segments": [
            {"name": name,
             "quarters": {f"FY{str(fy)[2:]} Q{fq}": val
                          for (fy, fq), val in sorted(qs.items())}}
            for name, qs in merged.items()],
    }


def ingest(url: str) -> dict:
    wb = download_workbook(url)
    grids = workbook_to_grids(wb)
    return {
        "source_url": url,
        "fetched_at": datetime.now().isoformat(timespec="seconds"),
        "sheet_names": list(grids.keys()),
        "segments": extract_segments(grids),
        "sheets": grids,
    }


if __name__ == "__main__":
    import sys
    result = ingest(sys.argv[1])
    print(json.dumps({"sheet_names": result["sheet_names"],
                      "segments_found": len(result["segments"])}, indent=2))
